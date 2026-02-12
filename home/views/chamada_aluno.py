from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction, IntegrityError
from django.contrib.auth.decorators import login_required


from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from django.db.models import Count, Q
from django.utils import timezone
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib import colors
import csv
from openpyxl import Workbook

from home.models import Chamada

import io
import json
from datetime import datetime

from home.models import (
    Turma,
    Disciplina,
    Docente,
    TurmaDisciplina,
    Chamada,
    Presenca,
    Aluno,
    DiarioDeClasse
)

# ======================================================
# Funções auxiliares
# ======================================================

def user_has_role(user, roles):
    if isinstance(roles, str):
        roles = [r.strip() for r in roles.split(",")]
    return hasattr(user, "role") and user.role in roles


def get_professor_or_gestor(user):
    professor = Docente.objects.filter(user=user).first()
    if professor:
        return professor

    if user_has_role(user, ["diretor", "coordenador"]):
        return None

    return "bloqueado"


# ======================================================
# 1) TELA PRINCIPAL DE CHAMADA
# ======================================================
@login_required
def tela_chamada(request):
    user = request.user

    # -------------------------------------------------
    # Papéis permitidos
    # -------------------------------------------------
    roles_permitidos = ["professor", "coordenador", "diretor"]

    if user.role not in roles_permitidos:
        return HttpResponseForbidden("Acesso negado.")

    hoje = datetime.now().date().strftime("%Y-%m-%d")

    # -------------------------------------------------
    # PROFESSOR
    # -------------------------------------------------
    if user.role == "professor":
        prof_obj = Docente.objects.filter(
            user=user,
            escola=user.escola
        ).first()

        if not prof_obj:
            return HttpResponseForbidden("Professor sem vínculo docente.")

        turmas_disciplinas = (
            TurmaDisciplina.objects
            .filter(
                professor=prof_obj,
                escola=user.escola
            )
            .select_related("turma", "disciplina")
        )

    # -------------------------------------------------
    # COORDENADOR / DIRETOR
    # -------------------------------------------------
    else:
        turmas_disciplinas = (
            TurmaDisciplina.objects
            .filter(escola=user.escola)
            .select_related("turma", "disciplina")
        )

    # -------------------------------------------------
    # Extrair listas sem duplicação
    # -------------------------------------------------
    turmas = sorted(
        {td.turma for td in turmas_disciplinas},
        key=lambda t: t.nome
    )

    disciplinas = sorted(
        {td.disciplina for td in turmas_disciplinas},
        key=lambda d: d.nome
    )

    return render(
        request,
        "pages/chamada/realizar_chamadas.html",
        {
            "turmas": turmas,
            "disciplinas": disciplinas,
            "data_hoje": hoje,
        }
    )

# ======================================================
# 2) API – CARREGAR ALUNOS DA TURMA
# ======================================================
@login_required
def api_carregar_alunos(request, turma_id):

    try:
        turma = Turma.objects.get(id=turma_id, escola=request.user.escola)
    except Turma.DoesNotExist:
        return JsonResponse({"erro": "Turma não encontrada."}, status=404)

    alunos = (
        turma.alunos.filter(ativo=True)
        .order_by("nome")
        .values("id", "nome")
    )

    return JsonResponse({"alunos": list(alunos)})


# ======================================================
# 3) SALVAR PRESENÇAS
# ======================================================
@csrf_exempt
@login_required
def salvar_presencas(request):

    acesso = get_professor_or_gestor(request.user)
    if acesso == "bloqueado":
        return JsonResponse(
            {"status": "erro", "mensagem": "Acesso negado."},
            status=403
        )

    if request.method != "POST":
        return JsonResponse(
            {"status": "erro", "mensagem": "Método não permitido"},
            status=405
        )

    try:
        data = json.loads(request.body)
    except Exception:
        return JsonResponse(
            {"status": "erro", "mensagem": "JSON inválido"},
            status=400
        )

    turma_id = data.get("turma")
    disciplina_id = data.get("disciplina")
    data_aula = data.get("data")
    lista = data.get("lista", [])

    if not turma_id or not disciplina_id or not data_aula:
        return JsonResponse(
            {"status": "erro", "mensagem": "Campos obrigatórios faltando."},
            status=400
        )

    # ===============================
    # PROFESSOR (SE FOR PERFIL PROFESSOR)
    # ===============================
    professor = None

    if acesso == "professor":
        professor = Docente.objects.filter(
            user=request.user,
            escola=request.user.escola
        ).first()

        if not professor:
            return JsonResponse(
                {
                    "status": "erro",
                    "mensagem": "Professor não está vinculado corretamente."
                },
                status=400
            )

    # ===============================
    # TURMA
    # ===============================
    try:
        turma = Turma.objects.get(
            id=turma_id,
            escola=request.user.escola
        )
    except Turma.DoesNotExist:
        return JsonResponse(
            {"status": "erro", "mensagem": "Turma inválida."},
            status=404
        )

    # ===============================
    # DISCIPLINA
    # ===============================
    try:
        disciplina = Disciplina.objects.get(
            id=disciplina_id,
            escola=request.user.escola
        )
    except Disciplina.DoesNotExist:
        return JsonResponse(
            {"status": "erro", "mensagem": "Disciplina inválida."},
            status=404
        )

    # ===============================
    # VALIDAÇÃO DE VÍNCULO (SÓ PROFESSOR)
    # ===============================
    if acesso == "professor":
        if not TurmaDisciplina.objects.filter(
            turma=turma,
            disciplina=disciplina,
            professor=professor
        ).exists():
            return JsonResponse(
                {
                    "status": "erro",
                    "mensagem": "Você não está vinculado a esta disciplina nesta turma."
                },
                status=403
            )

    # ===============================
    # DATA
    # ===============================
    try:
        data_aula = datetime.strptime(data_aula, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse(
            {"status": "erro", "mensagem": "Data inválida."},
            status=400
        )

    erros_alunos = []

    # ===============================
    # TRANSAÇÃO
    # ===============================
    try:
        with transaction.atomic():

            # --------------------------------
            # DIÁRIO DE CLASSE (ENTIDADE RAIZ)
            # --------------------------------
            diario, _ = DiarioDeClasse.objects.get_or_create(
                data_ministrada=data_aula,
                turma=turma,
                disciplina=disciplina,
                professor=professor if acesso == "professor" else None,
                escola=turma.escola,
                defaults={
                    "criado_por": request.user,
                    "status": "REALIZADA"
                }
            )

            # --------------------------------
            # CHAMADA (UMA POR DIÁRIO)
            # --------------------------------
            chamada, _ = Chamada.objects.get_or_create(
                diario=diario,
                defaults={
                    "criado_por": request.user
                }
            )

            # --------------------------------
            # PRESENÇAS (UMA POR ALUNO)
            # --------------------------------
            for item in lista:
                aluno_id = item.get("aluno_id")
                presente = item.get("presente", False)
                obs = item.get("observacao", "")

                try:
                    aluno = Aluno.objects.get(
                        id=aluno_id,
                        escola=request.user.escola
                    )
                except Aluno.DoesNotExist:
                    erros_alunos.append({
                        "aluno_id": aluno_id,
                        "mensagem": "Aluno não encontrado."
                    })
                    continue

                Presenca.objects.update_or_create(
                    chamada=chamada,
                    aluno=aluno,
                    defaults={
                        "presente": presente,
                        "observacao": obs
                    }
                )

    except IntegrityError:
        return JsonResponse(
            {
                "status": "erro",
                "mensagem": "Erro de integridade ao salvar a chamada."
            },
            status=400
        )

    # ===============================
    # RETORNO FINAL
    # ===============================
    if erros_alunos:
        return JsonResponse(
            {"status": "parcial", "erros": erros_alunos},
            status=207
        )

    return JsonResponse(
        {
            "status": "sucesso",
            "mensagem": "Chamada salva com sucesso!"
        }
    )


@login_required
def disciplinas_por_turma(request, turma_id):
    turma = get_object_or_404(
        Turma,
        id=turma_id,
        escola=request.user.escola
    )

    qs = TurmaDisciplina.objects.filter(
        turma=turma
    ).select_related("disciplina")

    disciplinas = [
        {
            "id": td.disciplina.id,
            "nome": td.disciplina.nome
        }
        for td in qs
    ]

    return JsonResponse(disciplinas, safe=False)



# ======================================================
# 4) HISTÓRICO DE CHAMADAS
# ======================================================
@login_required
def listar_chamadas(request):

    user = request.user
    hoje = datetime.now().date()
    hoje_str = hoje.strftime("%Y-%m-%d")

    filtro_data = request.GET.get("data")
    filtro_turma = request.GET.get("turma")
    filtro_disciplina = request.GET.get("disciplina")

    sem_filtros = not filtro_data and not filtro_turma and not filtro_disciplina

    professor = Docente.objects.filter(
        user=user,
        escola=user.escola
    ).first()

    # =====================================================
    # PERFIL PROFESSOR
    # =====================================================
    if professor:

        base = Chamada.objects.filter(
            diario__professor=professor,
            diario__turma__escola=user.escola
        )

        turmas = Turma.objects.filter(
            turmadisciplina__professor=professor,
            escola=user.escola
        ).distinct().order_by("nome")

        disciplinas = Disciplina.objects.filter(
            turmadisciplina__professor=professor,
            escola=user.escola
        ).distinct().order_by("nome")

    # =====================================================
    # PERFIL DIRETOR / COORDENADOR
    # =====================================================
    else:
        base = Chamada.objects.filter(
            diario__turma__escola=user.escola
        )

        turmas = Turma.objects.filter(
            escola=user.escola
        ).order_by("nome")

        disciplinas = Disciplina.objects.filter(
            escola=user.escola
        ).order_by("nome")

    # =====================================================
    # FILTROS
    # =====================================================
    if sem_filtros:
        base = base.filter(diario__data_ministrada=hoje)
        filtro_data = hoje_str

    if filtro_data:
        try:
            data_convertida = datetime.strptime(
                filtro_data, "%Y-%m-%d"
            ).date()
            base = base.filter(
                diario__data_ministrada=data_convertida
            )
        except ValueError:
            pass

    if filtro_turma:
        base = base.filter(diario__turma_id=filtro_turma)

    if filtro_disciplina:
        base = base.filter(diario__disciplina_id=filtro_disciplina)

    # =====================================================
    # QUERY FINAL
    # =====================================================
    chamadas_queryset = (
        base
        .select_related(
            "diario",
            "diario__turma",
            "diario__disciplina",
            "diario__professor",
        )
        .order_by(
            "-diario__data_ministrada",
            "diario__turma__nome",
            "diario__disciplina__nome",
        )
    )

    # =====================================================
    # PAGINAÇÃO
    # =====================================================
    from django.core.paginator import Paginator
    paginator = Paginator(chamadas_queryset, 20)
    pagina = request.GET.get("page")
    chamadas = paginator.get_page(pagina)

    # =====================================================
    # RENDER
    # =====================================================
    return render(
        request,
        "pages/chamada/listar_chamadas.html",
        {
            "chamadas": chamadas,
            "turmas": turmas,
            "disciplinas": disciplinas,
            "data_hoje": hoje_str,
            "filtro_data": filtro_data or "",
            "filtro_turma": filtro_turma or "",
            "filtro_disciplina": filtro_disciplina or "",
        }
    )


# ======================================================
# 5) DETALHE DA CHAMADA
# ======================================================
@login_required
def detalhe_chamada(request, chamada_id):

    acesso = get_professor_or_gestor(request.user)
    if acesso == "bloqueado":
        return render(request, "errors/403.html", status=403)

    chamada = get_object_or_404(
        Chamada.objects.select_related(
            "diario",
            "diario__turma",
            "diario__disciplina",
            "diario__professor",
        ),
        id=chamada_id,
        diario__turma__escola=request.user.escola
    )

    presencas = (
        Presenca.objects
        .filter(chamada=chamada)
        .select_related("aluno")
        .order_by("aluno__nome")
    )

    return render(
        request,
        "pages/chamada/detalhe_chamada.html",
        {
            "chamada": chamada,
            "diario": chamada.diario,
            "presencas": presencas,
        }
    )



# ======================================================
# 6) PDF DA CHAMADA
# ======================================================
@login_required
def pdf_chamada(request, chamada_id):

    acesso = get_professor_or_gestor(request.user)
    if acesso == "bloqueado":
        return render(request, "errors/403.html", status=403)

    chamada = get_object_or_404(
        Chamada,
        id=chamada_id,
        diario__turma__escola=request.user.escola
    )

    diario = chamada.diario

    presencas = (
        Presenca.objects
        .filter(chamada=chamada)
        .select_related("aluno")
        .order_by("aluno__nome")
    )

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)

    # =====================================================
    # CABEÇALHO
    # =====================================================
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(2 * cm, 28 * cm, "Registro de Chamada")

    pdf.setFont("Helvetica", 12)
    pdf.drawString(
        2 * cm, 26.8 * cm,
        f"Data: {diario.data_ministrada.strftime('%d/%m/%Y')}"
    )
    pdf.drawString(
        2 * cm, 26.2 * cm,
        f"Turma: {diario.turma.nome}"
    )
    pdf.drawString(
        2 * cm, 25.6 * cm,
        f"Disciplina: {diario.disciplina.nome}"
    )
    pdf.drawString(
        2 * cm, 25.0 * cm,
        f"Professor: {diario.professor.nome if diario.professor else '---'}"
    )

    # =====================================================
    # TABELA
    # =====================================================
    y = 23.5 * cm

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawString(2 * cm, y, "Aluno")
    pdf.drawString(11 * cm, y, "Presente")
    pdf.drawString(14 * cm, y, "Observação")

    pdf.setFont("Helvetica", 11)
    y -= 0.7 * cm

    for p in presencas:
        pdf.drawString(2 * cm, y, p.aluno.nome[:35])
        pdf.drawString(11 * cm, y, "✔" if p.presente else "✘")
        pdf.drawString(14 * cm, y, (p.observacao or "")[:25])
        y -= 0.6 * cm

        if y < 2 * cm:
            pdf.showPage()
            pdf.setFont("Helvetica", 11)
            y = 28 * cm

    pdf.showPage()
    pdf.save()

    buffer.seek(0)
    return HttpResponse(buffer, content_type="application/pdf")


# ======================================================
# 7) EDITAR CHAMADA
# ======================================================
@login_required
def editar_chamada(request, chamada_id):

    acesso = get_professor_or_gestor(request.user)
    if acesso == "bloqueado":
        return render(request, "errors/403.html", status=403)

    chamada = get_object_or_404(
        Chamada,
        id=chamada_id,
        turma__escola=request.user.escola
    )

    presencas = (
        Presenca.objects.filter(chamada=chamada)
        .select_related("aluno")
        .order_by("aluno__nome")
    )

    return render(request, "pages/chamada/editar_chamada.html", {
        "chamada": chamada,
        "presencas": presencas
    })


# ======================================================
# 8) ATUALIZAR CHAMADA
# ======================================================
@login_required
def atualizar_chamada(request, chamada_id):

    acesso = get_professor_or_gestor(request.user)
    if acesso == "bloqueado":
        return JsonResponse(
            {"status": "erro", "mensagem": "Acesso negado."},
            status=403
        )

    if request.method != "POST":
        return JsonResponse(
            {"status": "erro", "mensagem": "Método inválido"},
            status=405
        )

    try:
        data = json.loads(request.body)
    except:
        return JsonResponse(
            {"status": "erro", "mensagem": "JSON inválido"},
            status=400
        )

    lista = data.get("lista", [])

    chamada = get_object_or_404(
        Chamada,
        id=chamada_id,
        diario__turma__escola=request.user.escola
    )

    erros = []

    try:
        with transaction.atomic():

            for item in lista:
                aluno_id = item.get("aluno_id")
                presente = item.get("presente", False)
                obs = item.get("observacao", "")

                try:
                    aluno = Aluno.objects.get(
                        id=aluno_id,
                        escola=request.user.escola
                    )
                except:
                    erros.append({
                        "aluno_id": aluno_id,
                        "mensagem": "Aluno não encontrado."
                    })
                    continue

                Presenca.objects.update_or_create(
                    chamada=chamada,
                    aluno=aluno,
                    defaults={
                        "presente": presente,
                        "observacao": obs,
                    }
                )

            chamada.criado_por = request.user
            chamada.save()

    except Exception as e:
        return JsonResponse(
            {
                "status": "erro",
                "mensagem": "Falha ao atualizar a chamada.",
                "detalhe": str(e)
            },
            status=400
        )

    if erros:
        return JsonResponse({"status": "parcial", "erros": erros})

    return JsonResponse(
        {"status": "sucesso", "mensagem": "Chamada atualizada com sucesso."}
    )



def relatorio_chamadas(request):
    hoje = timezone.now().date()

    mes = request.GET.get("mes")
    ano = request.GET.get("ano") or hoje.year
    turma_id = request.GET.get("turma")
    disciplina_id = request.GET.get("disciplina")

    # ===============================
    # BASE QUERYSET (OTIMIZADO)
    # ===============================
    chamadas = (
        Chamada.objects
        .select_related(
            "diario",
            "diario__turma",
            "diario__disciplina",
            "diario__professor",
        )
        .annotate(
            presentes=Count(
                "presenca",
                filter=Q(presenca__presente=True)
            ),
            ausentes=Count(
                "presenca",
                filter=Q(presenca__presente=False)
            ),
        )
        .order_by(
            "-diario__data_ministrada",
            "diario__hora_inicio",
        )
    )

    # ===============================
    # FILTROS
    # ===============================
    if ano:
        chamadas = chamadas.filter(diario__data_ministrada__year=ano)

    if mes:
        chamadas = chamadas.filter(diario__data_ministrada__month=mes)

    if turma_id:
        chamadas = chamadas.filter(diario__turma_id=turma_id)

    if disciplina_id:
        chamadas = chamadas.filter(diario__disciplina_id=disciplina_id)

    # ===============================
    # DADOS AUXILIARES (FILTROS)
    # ===============================
    turmas = Turma.objects.all().order_by("nome")
    disciplinas = Disciplina.objects.all().order_by("nome")

    meses = [
        {"valor": 1, "nome": "Janeiro"},
        {"valor": 2, "nome": "Fevereiro"},
        {"valor": 3, "nome": "Março"},
        {"valor": 4, "nome": "Abril"},
        {"valor": 5, "nome": "Maio"},
        {"valor": 6, "nome": "Junho"},
        {"valor": 7, "nome": "Julho"},
        {"valor": 8, "nome": "Agosto"},
        {"valor": 9, "nome": "Setembro"},
        {"valor": 10, "nome": "Outubro"},
        {"valor": 11, "nome": "Novembro"},
        {"valor": 12, "nome": "Dezembro"},
    ]

    context = {
        "chamadas": chamadas,
        "turmas": turmas,
        "disciplinas": disciplinas,
        "meses": meses,
        "ano_atual": hoje.year,
    }

    return render(request, "pages/chamada/relatorio_chamadas.html", context)


def relatorio_chamadas_pdf(request):
    hoje = timezone.now().date()

    mes = int(request.GET.get("mes", hoje.month))
    ano = int(request.GET.get("ano", hoje.year))

    chamadas = (
        Chamada.objects
        .select_related(
            "diario",
            "diario__turma",
            "diario__disciplina",
            "diario__professor",
            "diario__escola",
        )
        .filter(
            diario__data_ministrada__year=ano,
            diario__data_ministrada__month=mes,
        )
        .annotate(
            presentes=Count("presenca", filter=Q(presenca__presente=True)),
            ausentes=Count("presenca", filter=Q(presenca__presente=False)),
        )
        .order_by("diario__data_ministrada")
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'inline; filename="relatorio_chamadas_{mes}_{ano}.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    elementos = []

    escola = chamadas.first().diario.escola if chamadas.exists() else None

    # ===============================
    # CABEÇALHO
    # ===============================
    elementos.append(
        Paragraph("<b>RELATÓRIO MENSAL DE CHAMADAS</b>", styles["Title"])
    )
    elementos.append(Spacer(1, 12))

    if escola:
        elementos.append(
            Paragraph(f"<b>Escola:</b> {escola.nome}", styles["Normal"])
        )

    elementos.append(
        Paragraph(
            f"<b>Período:</b> {mes:02d}/{ano}",
            styles["Normal"],
        )
    )
    elementos.append(Spacer(1, 20))

    # ===============================
    # TABELA
    # ===============================
    dados = [
        [
            "Data",
            "Turma",
            "Disciplina",
            "Professor",
            "Presentes",
            "Ausentes",
        ]
    ]

    total_presentes = 0
    total_ausentes = 0

    for chamada in chamadas:
        dados.append([
            chamada.diario.data_ministrada.strftime("%d/%m/%Y"),
            chamada.diario.turma.nome,
            chamada.diario.disciplina.nome,
            chamada.diario.professor.nome if chamada.diario.professor else "-",
            chamada.presentes,
            chamada.ausentes,
        ])
        total_presentes += chamada.presentes
        total_ausentes += chamada.ausentes

    # Linha de totais
    dados.append([
        "",
        "",
        "",
        "TOTAL",
        total_presentes,
        total_ausentes,
    ])

    tabela = Table(dados, repeatRows=1)

    tabela.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
        ("ALIGN", (4, 1), (-1, -1), "CENTER"),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONT", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))

    elementos.append(tabela)

    doc.build(elementos)
    return response


def resumo_mensal_turma_professor(request):
    hoje = timezone.now().date()

    mes = int(request.GET.get("mes", hoje.month))
    ano = int(request.GET.get("ano", hoje.year))

    # ===============================
    # BASE QUERYSET
    # ===============================
    resumo = (
        Chamada.objects
        .select_related(
            "diario",
            "diario__turma",
            "diario__professor",
        )
        .filter(
            diario__data_ministrada__year=ano,
            diario__data_ministrada__month=mes,
        )
        .values(
            "diario__turma__id",
            "diario__turma__nome",
            "diario__professor__id",
            "diario__professor__nome",
        )
        .annotate(
            total_aulas=Count("id", distinct=True),
            total_presentes=Count(
                "presenca",
                filter=Q(presenca__presente=True)
            ),
            total_ausentes=Count(
                "presenca",
                filter=Q(presenca__presente=False)
            ),
        )
        .order_by(
            "diario__turma__nome",
            "diario__professor__nome",
        )
    )

    meses = [
        {"valor": 1, "nome": "Janeiro"},
        {"valor": 2, "nome": "Fevereiro"},
        {"valor": 3, "nome": "Março"},
        {"valor": 4, "nome": "Abril"},
        {"valor": 5, "nome": "Maio"},
        {"valor": 6, "nome": "Junho"},
        {"valor": 7, "nome": "Julho"},
        {"valor": 8, "nome": "Agosto"},
        {"valor": 9, "nome": "Setembro"},
        {"valor": 10, "nome": "Outubro"},
        {"valor": 11, "nome": "Novembro"},
        {"valor": 12, "nome": "Dezembro"},
    ]

    context = {
        "resumo": resumo,
        "meses": meses,
        "mes_atual": mes,
        "ano_atual": ano,
    }

    return render(
        request,
        "pages/chamada/resumo_mensal_turma_professor.html",
        context
    )

def export_resumo_mensal_csv(request):
    hoje = timezone.now().date()

    mes = int(request.GET.get("mes", hoje.month))
    ano = int(request.GET.get("ano", hoje.year))

    resumo = (
        Chamada.objects
        .select_related(
            "diario",
            "diario__turma",
            "diario__professor",
        )
        .filter(
            diario__data_ministrada__year=ano,
            diario__data_ministrada__month=mes,
        )
        .values(
            "diario__turma__nome",
            "diario__professor__nome",
        )
        .annotate(
            total_aulas=Count("id", distinct=True),
            total_presentes=Count(
                "presenca",
                filter=Q(presenca__presente=True)
            ),
            total_ausentes=Count(
                "presenca",
                filter=Q(presenca__presente=False)
            ),
        )
        .order_by(
            "diario__turma__nome",
            "diario__professor__nome",
        )
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="resumo_mensal_{mes:02d}_{ano}.csv"'
    )

    writer = csv.writer(response)
    writer.writerow([
        "Turma",
        "Professor",
        "Total de Aulas",
        "Total de Presentes",
        "Total de Ausentes",
    ])

    for item in resumo:
        writer.writerow([
            item["diario__turma__nome"],
            item["diario__professor__nome"] or "-",
            item["total_aulas"],
            item["total_presentes"],
            item["total_ausentes"],
        ])

    return response


def export_resumo_mensal_excel(request):
    hoje = timezone.now().date()

    mes = int(request.GET.get("mes", hoje.month))
    ano = int(request.GET.get("ano", hoje.year))

    resumo = (
        Chamada.objects
        .select_related(
            "diario",
            "diario__turma",
            "diario__professor",
        )
        .filter(
            diario__data_ministrada__year=ano,
            diario__data_ministrada__month=mes,
        )
        .values(
            "diario__turma__nome",
            "diario__professor__nome",
        )
        .annotate(
            total_aulas=Count("id", distinct=True),
            total_presentes=Count(
                "presenca",
                filter=Q(presenca__presente=True)
            ),
            total_ausentes=Count(
                "presenca",
                filter=Q(presenca__presente=False)
            ),
        )
        .order_by(
            "diario__turma__nome",
            "diario__professor__nome",
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo Mensal"

    ws.append([
        "Turma",
        "Professor",
        "Total de Aulas",
        "Total de Presentes",
        "Total de Ausentes",
    ])

    for item in resumo:
        ws.append([
            item["diario__turma__nome"],
            item["diario__professor__nome"] or "-",
            item["total_aulas"],
            item["total_presentes"],
            item["total_ausentes"],
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="resumo_mensal_{mes:02d}_{ano}.xlsx"'
    )

    wb.save(response)
    return response


@login_required
def relatorio_anual_chamadas(request):

    acesso = get_professor_or_gestor(request.user)
    if acesso == "bloqueado":
        return render(request, "errors/403.html", status=403)

    ano = request.GET.get("ano")

    if not ano:
        ano = datetime.now().year
    else:
        ano = int(ano)

    resumo = (
        DiarioDeClasse.objects
        .filter(
            escola=request.user.escola,
            data_ministrada__year=ano
        )
        .values(
            "turma__nome",
            "disciplina__nome",
            "professor__nome",
        )
        .annotate(
            total_aulas=Count("id", distinct=True),
            total_presentes=Count(
                "chamada__presenca",
                filter=Q(chamada__presenca__presente=True),
                distinct=True
            ),
            total_ausentes=Count(
                "chamada__presenca",
                filter=Q(chamada__presenca__presente=False),
                distinct=True
            ),
        )
        .order_by("turma__nome", "disciplina__nome")
    )

    return render(
        request,
        "pages/chamada/relatorio_anual_chamadas.html",
        {
            "resumo": resumo,
            "ano": ano,
        }
    )


@login_required
def relatorio_anual_chamadas_pdf(request):

    acesso = get_professor_or_gestor(request.user)
    if acesso == "bloqueado":
        return render(request, "errors/403.html", status=403)

    ano = request.GET.get("ano")

    if not ano:
        return HttpResponse("Ano não informado.", status=400)

    ano = int(ano)

    resumo = (
        Presenca.objects
        .filter(
            chamada__diario__data_ministrada__year=ano,
            chamada__diario__turma__escola=request.user.escola
        )
        .values(
            "chamada__diario__turma__nome",
            "chamada__diario__disciplina__nome",
            "chamada__diario__professor__nome",
        )
        .annotate(
            total_aulas=Count("chamada", distinct=True),
            total_presentes=Count("id", filter=Q(presente=True)),
            total_ausentes=Count("id", filter=Q(presente=False)),
        )
        .order_by(
            "chamada__diario__turma__nome",
            "chamada__diario__disciplina__nome",
        )
    )

    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    y = height - 2 * cm

    # TÍTULO
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(2 * cm, y, "Relatório Anual de Chamadas")
    y -= 0.8 * cm

    pdf.setFont("Helvetica", 12)
    pdf.drawString(2 * cm, y, f"Ano: {ano}")
    y -= 0.5 * cm

    pdf.drawString(2 * cm, y, f"Escola: {request.user.escola.nome}")
    y -= 1.2 * cm

    # CABEÇALHO DA TABELA
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(2 * cm, y, "Turma")
    pdf.drawString(6 * cm, y, "Disciplina")
    pdf.drawString(10 * cm, y, "Professor")
    pdf.drawString(14 * cm, y, "Aulas")
    pdf.drawString(15.5 * cm, y, "P")
    pdf.drawString(17 * cm, y, "F")

    y -= 0.4 * cm
    pdf.line(2 * cm, y, 19 * cm, y)
    y -= 0.5 * cm

    pdf.setFont("Helvetica", 10)

    for r in resumo:
        if y < 2 * cm:
            pdf.showPage()
            y = height - 2 * cm
            pdf.setFont("Helvetica", 10)

        pdf.drawString(2 * cm, y, r["chamada__diario__turma__nome"][:20])
        pdf.drawString(6 * cm, y, r["chamada__diario__disciplina__nome"][:20])
        pdf.drawString(
            10 * cm,
            y,
            (r["chamada__diario__professor__nome"] or "—")[:18]
        )
        pdf.drawRightString(15 * cm, y, str(r["total_aulas"]))
        pdf.drawRightString(16.5 * cm, y, str(r["total_presentes"]))
        pdf.drawRightString(18 * cm, y, str(r["total_ausentes"]))

        y -= 0.45 * cm

    pdf.showPage()
    pdf.save()

    buffer.seek(0)
    return HttpResponse(buffer, content_type="application/pdf")


@login_required
def relatorio_anual_chamadas_excel(request):

    acesso = get_professor_or_gestor(request.user)
    if acesso == "bloqueado":
        return JsonResponse({"erro": "Acesso negado"}, status=403)

    ano = request.GET.get("ano")
    if not ano:
        return JsonResponse({"erro": "Ano não informado"}, status=400)

    ano = int(ano)

    resumo = (
        Presenca.objects
        .filter(
            chamada__diario__data_ministrada__year=ano,
            chamada__diario__turma__escola=request.user.escola
        )
        .values(
            "chamada__diario__turma__nome",
            "chamada__diario__disciplina__nome",
            "chamada__diario__professor__nome",
        )
        .annotate(
            total_aulas=Count("chamada", distinct=True),
            total_presentes=Count("id", filter=Q(presente=True)),
            total_ausentes=Count("id", filter=Q(presente=False)),
        )
        .order_by(
            "chamada__diario__turma__nome",
            "chamada__diario__disciplina__nome",
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Chamadas {ano}"

    # CABEÇALHO
    headers = [
        "Turma",
        "Disciplina",
        "Professor",
        "Total de Aulas",
        "Presentes",
        "Ausentes",
    ]

    ws.append(headers)

    # ESTILO DO CABEÇALHO
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # DADOS
    for r in resumo:
        ws.append([
            r["chamada__diario__turma__nome"],
            r["chamada__diario__disciplina__nome"],
            r["chamada__diario__professor__nome"] or "-",
            r["total_aulas"],
            r["total_presentes"],
            r["total_ausentes"],
        ])

    # AUTO AJUSTE DE COLUNAS
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="relatorio_chamadas_{ano}.xlsx"'
    )

    wb.save(response)
    return response



